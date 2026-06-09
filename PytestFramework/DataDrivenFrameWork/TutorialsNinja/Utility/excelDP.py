from openpyxl import Workbook
import openpyxl


def get_Excel_Data(path, sheet_name):
    fl = []
    Workbook = openpyxl.load_workbook(path)
    sheet = Workbook[sheet_name]
    tr = sheet.max_row
    tc = sheet.max_column

    for r in range(2, tr + 1):
        rl = []
        for c in range(1, tc + 1):
            rl.append(sheet.cell(r, c).value)
        fl.append(rl)
    return fl
